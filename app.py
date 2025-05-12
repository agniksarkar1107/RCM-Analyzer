import sys
import sqlite3
import streamlit as st
import os
from dotenv import load_dotenv
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from utils.document_processor import process_document
from utils.gemini import initialize_gemini, analyze_risk_with_gemini
from utils.db import initialize_chroma, store_in_chroma, query_chroma
import time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Check SQLite version and warn if incompatible with ChromaDB
sqlite_version = sqlite3.sqlite_version_info
min_sqlite_version = (3, 35, 0)
is_sqlite_compatible = sqlite_version >= min_sqlite_version

# Load environment variables
load_dotenv()

# Set API key directly as fallback if not loaded from .env
if not os.environ.get("GEMINI_API_KEY"):
    os.environ["GEMINI_API_KEY"] = "AIzaSyBdz-qcLFRDsR-mm37AlRf2w6RZws2lDL0"

# Set page configuration
st.set_page_config(
    page_title="Risk Control Matrix Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize Gemini
gemini_model = initialize_gemini()

def main():
    # Add custom CSS
    st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        color: #1E88E5;
        text-align: center;
    }
    .sub-header {
        font-size: 1.5rem;
        color: #424242;
        margin-bottom: 20px;
    }
    .card {
        padding: 20px;
        border-radius: 5px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin-bottom: 20px;
    }
    .dept-card {
        padding: 15px;
        border-radius: 5px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        margin-bottom: 15px;
        background-color: transparent;
    }
    .high-risk {
        /* Removed border-left */
    }
    .medium-risk {
        /* Removed border-left */
    }
    .low-risk {
        /* Removed border-left */
    }
    .analyzing {
        text-align: center;
        padding: 20px;
        border-radius: 5px;
        background-color: transparent;
        margin-bottom: 20px;
        font-size: 1.2rem;
        color: #2E7D32;
    }
    .stTabs [data-baseweb="tab-panel"] {
        padding-top: 0 !important;
    }
    /* Make streamlit components background transparent */
    .stExpander {
        background-color: transparent !important;
    }
    /* Make general streamlit elements background transparent */
    .stMarkdown, .stInfo {
        background-color: transparent !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Header
    st.markdown("<h1 class='main-header'>Risk Control Matrix Analyzer</h1>", unsafe_allow_html=True)
    st.markdown("<p class='sub-header'>Upload your RCM document for AI risk analysis</p>", unsafe_allow_html=True)
    
    # Display SQLite compatibility warning if needed
    if not is_sqlite_compatible:
        st.warning(f"Warning: Your system has SQLite version {'.'.join(map(str, sqlite_version))}, which is below the recommended version {'.'.join(map(str, min_sqlite_version))} for ChromaDB persistence. Vector storage will use in-memory mode.")
    
    # Create a simple upload interface
    uploaded_file = st.file_uploader("Upload Risk Control Matrix document", 
                                   type=["xlsx", "csv", "pdf", "docx"], 
                                   help="Upload RCM file to analyze")
    
    if uploaded_file:
        col1, col2 = st.columns([1, 3])
        
        with col1:
            st.info(f"Uploaded: {uploaded_file.name}")
            analyze_button = st.button("Analyze Document", type="primary")
        
        with col2:
            st.info("This tool will analyze your Risk Control Matrix and identify key risks across departments.")
    
    # Main content area
    if 'analyzed_data' not in st.session_state:
        st.session_state.analyzed_data = None
        
    if uploaded_file and analyze_button:
        with st.spinner():
            # Show "being analyzed by AI" message
            st.markdown("<div class='analyzing'><b>🔍 Being analyzed by AI...</b><br>Examining departments, control objectives, and identifying risk patterns</div>", unsafe_allow_html=True)
            
            # Save the uploaded file temporarily
            temp_file_path = f"temp_{uploaded_file.name}"
            with open(temp_file_path, "wb") as f:
                f.write(uploaded_file.getvalue())
            
            # Process document
            try:
                processed_data = process_document(temp_file_path)
                
                # Try to store in ChromaDB, but continue if it fails
                try:
                    if is_sqlite_compatible:
                        db = initialize_chroma("risk_control_matrix")
                        store_in_chroma(db, processed_data)
                    else:
                        # Skip ChromaDB storage if SQLite is incompatible but don't show error
                        pass
                except Exception as chroma_error:
                    # Log the error but don't display to user unless debugging
                    print(f"ChromaDB storage failed: {str(chroma_error)}")
                
                # Analyze with Gemini
                st.session_state.analyzed_data = analyze_risk_with_gemini(
                    gemini_model,
                    processed_data
                )
                
                # Try to remove temp file, but don't fail if it can't be removed
                try:
                    os.remove(temp_file_path)
                except PermissionError:
                    # Log the error but continue execution
                    print(f"Could not remove temporary file - it will be cleaned up later.")
                
                st.success("Analysis complete!")
                time.sleep(1)
                st.rerun()
                
            except Exception as e:
                st.error(f"Error analyzing document: {str(e)}")
                # Try to remove temp file, but don't fail if it can't be removed
                try:
                    if os.path.exists(temp_file_path):
                        os.remove(temp_file_path)
                except PermissionError:
                    # Log the error but continue execution
                    pass
    
    # Display analyzed data if available
    if st.session_state.analyzed_data:
        display_simplified_analysis(st.session_state.analyzed_data)

def create_downloadable_excel(data):
    """
    Create an Excel file with all analyzed data and formatting
    
    Args:
        data: The analyzed data from Gemini
        
    Returns:
        Excel file as bytes
    """
    # Create workbook
    wb = Workbook()
    
    # Define styles
    title_font = Font(name='Arial', size=14, bold=True, color="0000FF")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    risk_high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    risk_medium_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    risk_low_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Risk Summary"
    
    # Add title
    ws_summary.cell(row=1, column=1, value="Risk Control Matrix Analysis Summary").font = title_font
    ws_summary.merge_cells('A1:H1')
    
    # Add summary info
    ws_summary.cell(row=3, column=1, value="Departments").font = header_font
    ws_summary.cell(row=3, column=2, value=", ".join(data.get("departments", [])))
    
    # Risk distribution
    risk_dist = data.get("risk_distribution", {})
    ws_summary.cell(row=4, column=1, value="Risk Distribution").font = header_font
    ws_summary.cell(row=5, column=1, value="High Risk Items")
    ws_summary.cell(row=5, column=2, value=risk_dist.get("High", 0))
    ws_summary.cell(row=5, column=2).fill = risk_high_fill
    
    ws_summary.cell(row=6, column=1, value="Medium Risk Items")
    ws_summary.cell(row=6, column=2, value=risk_dist.get("Medium", 0))
    ws_summary.cell(row=6, column=2).fill = risk_medium_fill
    
    ws_summary.cell(row=7, column=1, value="Low Risk Items")
    ws_summary.cell(row=7, column=2, value=risk_dist.get("Low", 0))
    ws_summary.cell(row=7, column=2).fill = risk_low_fill
    
    # Control Objectives sheet
    ws_controls = wb.create_sheet(title="Control Objectives")
    
    # Headers - Updated with all columns from the image
    headers = [
        "Department", "Control Objective", "What Can Go Wrong", "Risk Level", 
        "Control Activities", "Person(s) in charge of existing control", "Additional Remarks (if any)",
        "Risk of material misstatement - Key Control - Yes/No", "Frequency Control",
        "Balance Sheet", "P&L", "Automated/Manual", "Preventive/Detective",
        "Existence/occurrence", "Completeness", "Valuation/Accuracy",
        "Rights/Obligations", "Presentation/Disclosure", "Cut-off",
        "Control/Design Gap", "Proposed Solution"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_controls.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    # Add control objectives data
    control_objectives = data.get("control_objectives", [])
    for i, obj in enumerate(control_objectives, start=2):
        ws_controls.cell(row=i, column=1, value=obj.get("department", ""))
        ws_controls.cell(row=i, column=2, value=obj.get("objective", ""))
        ws_controls.cell(row=i, column=3, value=obj.get("what_can_go_wrong", ""))
        
        risk_level = obj.get("risk_level", "Medium")
        ws_controls.cell(row=i, column=4, value=risk_level)
        
        # Color based on risk level
        if risk_level == "High":
            ws_controls.cell(row=i, column=4).fill = risk_high_fill
        elif risk_level == "Medium":
            ws_controls.cell(row=i, column=4).fill = risk_medium_fill
        elif risk_level == "Low":
            ws_controls.cell(row=i, column=4).fill = risk_low_fill
            
        # Control Activities - Generate detailed content if empty
        control_activities = obj.get("control_activities", "")
        if not control_activities:
            # Generate detailed control activities based on the risk
            what_can_go_wrong = obj.get("what_can_go_wrong", "")
            if "unauthorized access" in what_can_go_wrong.lower():
                control_activities = "Implementation of role-based access controls with regular access reviews. Multi-factor authentication for critical systems. Automated logging and monitoring of all access attempts. Regular audit of user privileges to ensure principle of least privilege."
            elif "database" in what_can_go_wrong.lower():
                control_activities = "Regular database health monitoring with automated alerts. Scheduled database integrity checks and maintenance. Comprehensive backup procedures with regular recovery testing. Database access strictly controlled through application interfaces only."
            elif "accounting entries" in what_can_go_wrong.lower() or "financial" in what_can_go_wrong.lower():
                control_activities = "Multi-level approval workflow for all journal entries. Automated validation of accounting codes and amounts. Regular reconciliation of accounts. Monthly review of unusual transactions and threshold-based exception reporting."
            else:
                control_activities = "Regular monitoring and review of processes. Clearly documented procedures with designated responsibilities. Automated controls where possible, with manual oversight. Periodic testing and validation of control effectiveness."
        
        ws_controls.cell(row=i, column=5, value=control_activities)
            
        # Add dummy data or actual data for additional columns
        ws_controls.cell(row=i, column=6, value=obj.get("person_in_charge", "Finance Manager"))  # Person in charge
        ws_controls.cell(row=i, column=7, value=obj.get("additional_remarks", ""))  # Additional Remarks
        ws_controls.cell(row=i, column=8, value=obj.get("key_control", "Yes"))  # Key Control
        ws_controls.cell(row=i, column=9, value=obj.get("frequency", "Monthly"))  # Frequency
        ws_controls.cell(row=i, column=10, value=obj.get("balance_sheet", "✓"))  # Balance Sheet
        ws_controls.cell(row=i, column=11, value=obj.get("p_l", "✓"))  # P&L
        ws_controls.cell(row=i, column=12, value=obj.get("automated_manual", "Manual"))  # Automated/Manual
        ws_controls.cell(row=i, column=13, value=obj.get("preventive_detective", "Preventive"))  # Preventive/Detective
        
        # Assertions
        ws_controls.cell(row=i, column=14, value=obj.get("existence_occurrence", "P"))  # Existence/occurrence
        ws_controls.cell(row=i, column=15, value=obj.get("completeness", "P"))  # Completeness
        ws_controls.cell(row=i, column=16, value=obj.get("valuation_accuracy", "P"))  # Valuation/Accuracy
        ws_controls.cell(row=i, column=17, value=obj.get("rights_obligations", ""))  # Rights/Obligations
        ws_controls.cell(row=i, column=18, value=obj.get("presentation_disclosure", ""))  # Presentation/Disclosure
        ws_controls.cell(row=i, column=19, value=obj.get("cut_off", ""))  # Cut-off
        
        # Modified: Control/Design Gap now uses "Yes" or "No" instead of gap details
        has_gap = "Yes" if obj.get("gap_details", "") else "No"
        ws_controls.cell(row=i, column=20, value=has_gap)
        
        # Use the proposed solution from the LLM directly
        proposed_solution = generate_proposed_solution(obj)
        ws_controls.cell(row=i, column=21, value=proposed_solution)
        
        # Apply borders
        for col in range(1, len(headers) + 1):
            ws_controls.cell(row=i, column=col).border = border
            ws_controls.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)
    
    # Add drop-down for Risk Level
    # Set a data validation list in column D from row 2 to last row with data
    dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    dv.add(f'D2:D{len(control_objectives)+1}')
    ws_controls.add_data_validation(dv)
    
    # Add drop-down for Yes/No fields
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    yes_no_dv.add(f'G2:G{len(control_objectives)+1}')  # Key Control column
    yes_no_dv.add(f'S2:S{len(control_objectives)+1}')  # Control/Design Gap column
    ws_controls.add_data_validation(yes_no_dv)
    
    # Add drop-down for Automated/Manual
    auto_manual_dv = DataValidation(type="list", formula1='"Automated,Manual,Automated/Manual"', allow_blank=True)
    auto_manual_dv.add(f'K2:K{len(control_objectives)+1}')
    ws_controls.add_data_validation(auto_manual_dv)
    
    # Add drop-down for Preventive/Detective
    prev_detect_dv = DataValidation(type="list", formula1='"Preventive,Detective,Preventive/Detective"', allow_blank=True)
    prev_detect_dv.add(f'L2:L{len(control_objectives)+1}')
    ws_controls.add_data_validation(prev_detect_dv)
    
    # Add drop-down for Frequency
    freq_dv = DataValidation(type="list", formula1='"Daily,Weekly,Monthly,Quarterly,Annually"', allow_blank=True)
    freq_dv.add(f'H2:H{len(control_objectives)+1}')
    ws_controls.add_data_validation(freq_dv)
    
    # Department Risk Analysis sheet
    ws_dept_risk = wb.create_sheet(title="Department Risk Analysis")
    
    # Headers
    dept_headers = ["Department", "Financial Risk", "Operational Risk", "Compliance Risk", 
                    "Strategic Risk", "Technological Risk", "Overall Risk"]
    
    for col, header in enumerate(dept_headers, start=1):
        cell = ws_dept_risk.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add department risk data
    dept_risks = data.get("department_risks", {})
    row = 2
    for dept, risks in dept_risks.items():
        ws_dept_risk.cell(row=row, column=1, value=dept)
        
        # Calculate overall risk for each department
        risk_values = []
        col = 2
        
        for risk_type in ["Financial", "Operational", "Compliance", "Strategic", "Technological"]:
            risk_value = risks.get(risk_type, 0)
            risk_values.append(risk_value)
            
            # Convert numeric value to text
            risk_text = "Low"
            risk_fill = risk_low_fill
            
            if risk_value >= 4:
                risk_text = "High"
                risk_fill = risk_high_fill
            elif risk_value >= 2:
                risk_text = "Medium"
                risk_fill = risk_medium_fill
                
            ws_dept_risk.cell(row=row, column=col, value=risk_text)
            ws_dept_risk.cell(row=row, column=col).fill = risk_fill
            ws_dept_risk.cell(row=row, column=col).border = border
            col += 1
        
        # Calculate overall risk
        avg_risk = sum(risk_values) / len(risk_values) if risk_values else 0
        risk_text = "Low"
        risk_fill = risk_low_fill
        
        if avg_risk >= 3.5:
            risk_text = "High"
            risk_fill = risk_high_fill
        elif avg_risk >= 2.0:
            risk_text = "Medium"
            risk_fill = risk_medium_fill
            
        ws_dept_risk.cell(row=row, column=col, value=risk_text)
        ws_dept_risk.cell(row=row, column=col).fill = risk_fill
        ws_dept_risk.cell(row=row, column=col).border = border
        
        row += 1
    
    # Recommendations sheet
    ws_recommendations = wb.create_sheet(title="Recommendations")
    
    # Headers
    rec_headers = ["Department", "Recommendation", "Priority", "Expected Impact"]
    
    for col, header in enumerate(rec_headers, start=1):
        cell = ws_recommendations.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add recommendations data
    recommendations = data.get("recommendations", [])
    for i, rec in enumerate(recommendations, start=2):
        ws_recommendations.cell(row=i, column=1, value=rec.get("department", ""))
        ws_recommendations.cell(row=i, column=2, value=rec.get("description", ""))
        
        priority = rec.get("priority", "Medium")
        ws_recommendations.cell(row=i, column=3, value=priority)
        
        # Color based on priority
        if priority == "High":
            ws_recommendations.cell(row=i, column=3).fill = risk_high_fill
        elif priority == "Medium":
            ws_recommendations.cell(row=i, column=3).fill = risk_medium_fill
        elif priority == "Low":
            ws_recommendations.cell(row=i, column=3).fill = risk_low_fill
            
        ws_recommendations.cell(row=i, column=4, value=rec.get("impact", ""))
        
        # Apply borders
        for col in range(1, 5):
            ws_recommendations.cell(row=i, column=col).border = border
            ws_recommendations.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)
    
    # Add drop-down for Priority
    dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    dv.add(f'C2:C{len(recommendations)+1}')
    ws_recommendations.add_data_validation(dv)
    
    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes

def generate_proposed_solution(obj):
    """Generate a proposed solution for an objective if one doesn't exist"""
    proposed_solution = obj.get("proposed_control", "")
    
    # Generate a fallback proposed solution if none exists
    if not proposed_solution:
        # Generate detailed proposed solution based on the risk
        what_can_go_wrong = obj.get("what_can_go_wrong", "").lower()
        if "unauthorized access" in what_can_go_wrong:
            proposed_solution = "Implement a comprehensive Identity and Access Management (IAM) solution with regular certification reviews. Establish segregation of duties matrix and enforce through automated controls. Implement privileged access management with just-in-time access."
        elif "database" in what_can_go_wrong:
            proposed_solution = "Implement database activity monitoring tools to track all changes. Establish formal change management procedures for schema and data modifications. Implement data loss prevention controls with automated alerting."
        elif "accounting" in what_can_go_wrong or "financial" in what_can_go_wrong:
            proposed_solution = "Implement automated validation rules for accounting entries with threshold-based approval workflows. Establish regular account reconciliation practices with management sign-off. Implement continuous monitoring dashboards for financial data integrity."
        else:
            proposed_solution = "Implement comprehensive documentation of control procedures with clear ownership. Establish regular control testing schedule with measurable effectiveness criteria. Enhance monitoring through automated dashboard reporting of control metrics."
    
    return proposed_solution

def display_simplified_analysis(data):
    """Display a simplified analysis focusing on departmental risks and gaps"""
    
    # Overview section
    st.markdown("## Departmental Risk Analysis")
    
    # Process departments
    departments = list(data.get("department_risks", {}).keys()) if data.get("department_risks") else data.get("departments", [])
    control_objectives = data.get("control_objectives", [])
    gaps = data.get("gaps", [])
    
    if not departments:
        st.warning("No departments found in the document. Please check the file format.")
        return
    
    # Add download button for Excel report
    excel_bytes = create_downloadable_excel(data)
    st.download_button(
        label="📥 Download Analysis as Excel",
        data=excel_bytes,
        file_name="risk_control_matrix_analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Download a fully formatted Excel file with all analysis data",
    )
    
    # Create CSV download button
    csv_data = io.StringIO()
    
    # Create CSV with control objectives data
    control_df = pd.DataFrame([
        {
            "Department": obj.get("department", ""),
            "Control Objective": obj.get("objective", ""),
            "What Can Go Wrong": obj.get("what_can_go_wrong", ""),
            "Risk Level": obj.get("risk_level", ""),
            "Control Activities": obj.get("control_activities", ""),
            "Person(s) in charge": obj.get("person_in_charge", "Finance Manager"),
            "Additional Remarks": obj.get("additional_remarks", ""),
            "Key Control": obj.get("key_control", "Yes"),
            "Frequency Control": obj.get("frequency", "Monthly"),
            "Balance Sheet": obj.get("balance_sheet", "✓"),
            "P&L": obj.get("p_l", "✓"),
            "Automated/Manual": obj.get("automated_manual", "Manual"),
            "Preventive/Detective": obj.get("preventive_detective", "Preventive"),
            "Existence/occurrence": obj.get("existence_occurrence", "P"),
            "Completeness": obj.get("completeness", "P"),
            "Valuation/Accuracy": obj.get("valuation_accuracy", "P"),
            "Rights/Obligations": obj.get("rights_obligations", ""),
            "Presentation/Disclosure": obj.get("presentation_disclosure", ""),
            "Cut-off": obj.get("cut_off", ""),
            "Control/Design Gap": "Yes" if obj.get("gap_details", "") else "No",
            "Proposed Solution": generate_proposed_solution(obj)
        }
        for obj in control_objectives
    ])
    
    control_df.to_csv(csv_data, index=False)
    
    st.download_button(
        label="📥 Download Analysis as CSV",
        data=csv_data.getvalue(),
        file_name="risk_control_matrix_analysis.csv",
        mime="text/csv",
        help="Download a CSV file with analysis data",
    )
    
    # Create tabs for each department
    dept_tabs = st.tabs(departments)
    
    # Create analysis for each department
    for i, dept in enumerate(departments):
        with dept_tabs[i]:
            # Get department-specific data
            dept_objectives = [obj for obj in control_objectives if obj.get("department") == dept]
            dept_gaps = [gap for gap in gaps if gap.get("department") == dept]
            
            # Department overview
            dept_risk_data = data.get("department_risks", {}).get(dept, {})
            if dept_risk_data:
                risk_level = dept_risk_data.get("overall_risk_level", "Medium")
                risk_class = "high-risk" if risk_level == "High" else "medium-risk" if risk_level == "Medium" else "low-risk"
                
                st.markdown(f"<div class='dept-card {risk_class}'>", unsafe_allow_html=True)
                st.markdown(f"### {dept} Department - {risk_level} Risk")
                st.markdown(f"**Summary**: {dept_risk_data.get('summary', 'No summary available.')}")
                st.markdown("</div>", unsafe_allow_html=True)
            
            # Risk Categories Analysis
            st.markdown("### Risk Type Analysis")
            
            # Define the specific risk types the user wants to analyze
            risk_types = ["Operational", "Financial", "Fraud", "Financial Fraud", "Operational Fraud"]
            
            # Check if we have RAG analysis results with risk_types directly
            if "risk_types" in dept_risk_data or "risk_analysis" in dept_risk_data:
                risk_type_data = dept_risk_data.get("risk_types", dept_risk_data.get("risk_analysis", {}))
                
                cols = st.columns(len(risk_types))
                for i, risk_type in enumerate(risk_types):
                    with cols[i]:
                        risks = risk_type_data.get(risk_type, [])
                        count = len(risks)
                        level = "High" if count >= 3 else "Medium" if count >= 1 else "Low"
                        color = "#FF5252" if level == "High" else "#FFC107" if level == "Medium" else "#4CAF50"
                        
                        st.markdown(f"<div style='text-align:center'>", unsafe_allow_html=True)
                        st.markdown(f"<h4>{risk_type}</h4>", unsafe_allow_html=True)
                        st.markdown(f"<h2 style='color:{color}'>{count}</h2>", unsafe_allow_html=True)
                        st.markdown(f"<span style='color:{color}'>{level} Risk</span>", unsafe_allow_html=True)
                        st.markdown(f"</div>", unsafe_allow_html=True)
                
                # Show specific risks for each type in an expander
                with st.expander("View Specific Risks by Type"):
                    for risk_type in risk_types:
                        risks = risk_type_data.get(risk_type, [])
                        if risks:
                            st.markdown(f"#### {risk_type} Risks")
                            for risk in risks:
                                st.markdown(f"- {risk}")
                            st.markdown("---")
            else:
                # Analyze content to detect risk types - fallback to keyword analysis
                risk_findings = {}
                for risk_type in risk_types:
                    risk_findings[risk_type] = []
                    
                    # Keywords for each risk type
                    keywords = {
                        "Operational": ["process", "workflow", "efficiency", "performance", "delivery", "resource", "procedure"],
                        "Financial": ["financial", "budget", "cost", "expense", "revenue", "payment", "accounting"],
                        "Fraud": ["fraud", "misappropriation", "theft", "falsification", "bribery", "corruption"],
                        "Financial Fraud": ["financial fraud", "embezzlement", "accounting fraud", "false reporting", "misstatement"],
                        "Operational Fraud": ["operational fraud", "process manipulation", "override", "unauthorized"]
                    }
                    
                    # Check each objective for this risk type
                    for obj in dept_objectives:
                        objective = obj.get("objective", "").lower()
                        risk = obj.get("what_can_go_wrong", "").lower()
                        
                        # Check if any keywords match
                        matched = False
                        for keyword in keywords.get(risk_type, []):
                            if keyword in objective or keyword in risk:
                                risk_findings[risk_type].append({
                                    "objective": obj.get("objective", ""),
                                    "risk": obj.get("what_can_go_wrong", ""),
                                    "risk_level": obj.get("risk_level", "Medium")
                                })
                                matched = True
                                break
                
                # Display risk findings
                cols = st.columns(len(risk_types))
                for i, risk_type in enumerate(risk_types):
                    with cols[i]:
                        count = len(risk_findings[risk_type])
                        level = "High" if count >= 3 else "Medium" if count >= 1 else "Low"
                        color = "#FF5252" if level == "High" else "#FFC107" if level == "Medium" else "#4CAF50"
                        
                        st.markdown(f"<div style='text-align:center'>", unsafe_allow_html=True)
                        st.markdown(f"<h4>{risk_type}</h4>", unsafe_allow_html=True)
                        st.markdown(f"<h2 style='color:{color}'>{count}</h2>", unsafe_allow_html=True)
                        st.markdown(f"<span style='color:{color}'>{level} Risk</span>", unsafe_allow_html=True)
                        st.markdown(f"</div>", unsafe_allow_html=True)
            
            # Control Gaps and Recommendations
            st.markdown("### Control Gaps and Recommendations")
            
            if dept_gaps:
                for i, gap in enumerate(dept_gaps):
                    st.markdown(f"<div class='dept-card high-risk'>", unsafe_allow_html=True)
                    st.markdown(f"**Gap {i+1}**: {gap.get('gap_title', '')}")
                    st.markdown(f"**Control Objective**: {gap.get('control_objective', '')}")
                    st.markdown(f"**Impact**: {gap.get('risk_impact', '')}")
                    
                    # Get or generate recommendation
                    recommendation = gap.get('proposed_solution', '')
                    if not recommendation:
                        # Find matching recommendation from overall recommendations
                        for rec in data.get("recommendations", []):
                            if rec.get("department") == dept:
                                recommendation = rec.get("description", "")
                                break
                    
                    st.markdown(f"**Recommended Action**: {recommendation}")
                    st.markdown("</div>", unsafe_allow_html=True)
            else:
                st.info("No control gaps identified for this department.")
            
            # Key Risks
            st.markdown("### Key Risks")
            key_risks = dept_risk_data.get("key_risks", [])
            if key_risks:
                for risk in key_risks:
                    st.markdown(f"- {risk}")
            else:
                # Generate a list of key risks if none exist
                if dept_objectives:
                    st.markdown("Based on analysis of control objectives:")
                    high_risks = [obj for obj in dept_objectives if obj.get("risk_level", "").lower() in ["high", "h", "critical"]]
                    for i, risk in enumerate(high_risks[:3]):  # Show up to 3 high risks
                        st.markdown(f"- {risk.get('what_can_go_wrong', '')}")
                else:
                    st.info("No key risks identified for this department.")
    
    # Overall Recommendations
    st.markdown("## Overall Recommendations")
    recommendations = data.get("recommendations", [])
    if recommendations:
        for i, rec in enumerate(recommendations[:5]):  # Show top 5 recommendations
            priority = rec.get("priority", "Medium")
            rec_class = "high-risk" if priority == "High" else "medium-risk" if priority == "Medium" else "low-risk"
            
            st.markdown(f"<div class='dept-card {rec_class}'>", unsafe_allow_html=True)
            st.markdown(f"**{rec.get('title', f'Recommendation {i+1}')}** (Priority: {priority})")
            st.markdown(f"{rec.get('description', '')}")
            
            if "impact" in rec:
                st.markdown(f"**Expected Impact**: {rec.get('impact', '')}")
                
            if "department" in rec and rec["department"]:
                st.markdown(f"**Department**: {rec.get('department', '')}")
                
            st.markdown("</div>", unsafe_allow_html=True)
    else:
        st.info("No specific recommendations generated. Consider reviewing the identified gaps.")

    # Add clear button at the bottom
    st.markdown("---")
    if st.button("🗑️ Clear Analysis", type="primary", help="Clear all analysis data and start fresh"):
        # Reset session state
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()

if __name__ == "__main__":
    main() 